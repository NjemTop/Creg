import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from django.conf import settings
from django.core.mail import EmailMessage
from pathlib import Path
import os
import logging
import traceback
from main.models import ServiseCard, SLAPolicy
from logger.log_config import setup_logger, get_abs_log_path


scripts_error_logger = setup_logger('scripts_error', get_abs_log_path('scripts_errors.log'), logging.ERROR)
scripts_info_logger = setup_logger('scripts_info', get_abs_log_path('scripts_info.log'), logging.INFO)


def format_duration(duration):
    """
    Преобразует длительность в формат 'дни часы минуты'.
    
    Аргументы:
    duration (datetime.timedelta): Длительность для преобразования.

    Возвращает:
    str: Длительность в формате 'дни часы минуты'.
    """
    if not duration:
        return "0д 0ч 0м"
    days = duration.days
    seconds = duration.seconds
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    return f"{days}д {hours}ч {minutes}м"

def prepare_report_data(data):
    """
    Подготавливает данные отчета для таблицы.

    Аргументы:
    data (QuerySet): Набор данных с информацией о тикетах.

    Возвращает:
    list: Список словарей с данными для отчета.
    """
    report_data = []
    sla_time_formatted = None
    try:
        for ticket in data:
            ### Преобразование времени SLA в дни, часы и минуты
            first_response_time_formatted = format_duration(ticket.first_response_time)

            if ticket.closed_date:
                sla_time_formatted = format_duration(ticket.sla_time)
            else:
                sla_time_formatted = "-"

            ### Формирование информации о нарушении SLA
            sla_violated = "Да" if ticket.sla_violated else "Нет"
            response_violated = "Да" if ticket.response_violated else "Нет"
            sla_overdue_formatted = format_duration(ticket.sla_overdue_time) if ticket.sla_overdue_time else ""
            response_overdue_formatted = format_duration(ticket.response_overdue_time) if ticket.response_overdue_time else ""

            if ticket.sla_violated and ticket.response_violated:
                sla_violation_info = f"Да - SLA: {sla_overdue_formatted}, Response: {response_overdue_formatted}"
            elif ticket.sla_violated:
                sla_violation_info = f"Да - {sla_overdue_formatted}"
            elif ticket.response_violated:
                sla_violation_info = f"Да - {response_overdue_formatted}"
            else:
                sla_violation_info = "Нет"

            ### Форматирование дат создания и закрытия заявки
            creation_date_formatted = ticket.creation_date.strftime('%d.%m.%Y %H:%M:%S')
            closed_date_formatted = ticket.closed_date.strftime('%d.%m.%Y %H:%M:%S') if ticket.closed_date else "-"

            ### Добавление данных в список отчетов
            report_data.append({
                'Номер': ticket.unique_id,
                'Тип': ticket.type_name,
                'Приоритет': ticket.priority,
                'Тема': ticket.subject,
                'Инициатор': ticket.initiator,
                'Исполнитель': ticket.assignee_name,
                'Статус': "В работе" if ticket.status != "Выполнено" else "Выполнено",
                'Дата создания': creation_date_formatted,
                'Дата решения': closed_date_formatted,
                'Первое время ответа': first_response_time_formatted,
                'Время решения': sla_time_formatted,
                'Нарушение SLA': sla_violation_info
            })

        ### Сортировка данных по дате создания
        report_data.sort(key=lambda x: datetime.strptime(x['Дата создания'], '%d.%m.%Y %H:%M:%S'))

    except Exception as e:
        scripts_error_logger.error(f"Ошибка при подготовке данных отчета: {traceback.format_exc()}")
    return report_data

def prepare_additional_info(client_name, start_date, end_date):
    """
    Подготавливает дополнительную информацию о клиенте и SLA.

    Аргументы:
    client_name (str): Название клиента.
    start_date (str): Дата начала отчетного периода.
    end_date (str): Дата окончания отчетного периода.

    Возвращает:
    dict: Словарь с дополнительной информацией о клиенте и SLA.
    """
    try:
        ### Получение информации о тарифном плане клиента
        servise_card = ServiseCard.objects.get(client_card__client_info__client_name=client_name)
        service_pack = servise_card.service_pack.lower()  # Преобразование к нижнему регистру

        ### Получение информации о плановом времени SLA для всех приоритетов
        sla_policies = SLAPolicy.objects.filter(plan=service_pack)

        sla_info = []
        for policy in sla_policies:
            sla_info.append({
                "Приоритет": policy.get_priority_display(),
                "Время реагирования": format_duration(policy.reaction_time),
                "Время разрешения (не более)": format_duration(policy.max_resolution_time)
            })

        if not sla_info:
            sla_info.append({
                "Приоритет": "Не указано",
                "Время реагирования": "Не указано",
                "Время разрешения (не более)": "Не указано"
            })

        ### Сортировка информации о SLA по приоритетам
        priority_order = ["Критический", "Высокий", "Средний", "Низкий"]
        sla_info.sort(key=lambda x: priority_order.index(x["Приоритет"]) if x["Приоритет"] in priority_order else len(priority_order))

        return {
            "Наименование клиента": client_name,
            "Период отчета": f"{start_date} - {end_date}",
            "Тарифный план": service_pack.capitalize(),
            "SLA Информация": sla_info
        }
    except ServiseCard.DoesNotExist:
        scripts_error_logger.error(f"Не удалось найти информацию об обслуживании клиента: {client_name}")
        return {
            "Наименование клиента": client_name,
            "Период отчета": f"{start_date} - {end_date}",
            "Тарифный план": "Не указан",
            "SLA Информация": [{
                "Приоритет": "Не указано",
                "Время реагирования": "Не указано",
                "Время разрешения (не более)": "Не указано"
            }]
        }
    except Exception as e:
        scripts_error_logger.error(f"Ошибка при подготовке информации о клиенте: {traceback.format_exc()}")
        return None

def create_excel_report(report_data, client_name, start_date, end_date):
    """
    Создает Excel отчет с данными и дополнительной информацией.

    Аргументы:
    report_data (list): Список словарей с данными для отчета.
    client_name (str): Название клиента.
    start_date (str): Дата начала отчетного периода.
    end_date (str): Дата окончания отчетного периода.

    Возвращает:
    str: Путь к созданному файлу отчета.
    """
    try:
        ### Получение текущей даты и формирование имени файла
        today = datetime.now()
        formatted_date = today.strftime("%d-%m-%Y")
        file_name = f"Отчет клиента {client_name} от {formatted_date}.xlsx"
        ### Путь до файла с отчётом
        file_path = Path(settings.MEDIA_ROOT) / 'reports' / file_name

        ### Создание DataFrame из данных отчета
        df = pd.DataFrame(report_data)
        additional_info = prepare_additional_info(client_name, start_date, end_date)

        if additional_info is None:
            raise ValueError("Не удалось подготовить дополнительную информацию для отчета.")

        ### Создание Excel файла и добавление форматирования
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            ### Запись данных отчета на первый лист
            df.to_excel(writer, index=False, sheet_name='Отчет')
            workbook = writer.book
            worksheet = writer.sheets['Отчет']
            customize_excel_sheet(worksheet)

            ### Создание второго листа с информацией о клиенте и SLA
            info_sla_worksheet = workbook.create_sheet(title='Информация и SLA')

            ### Заполнение информации о клиенте
            client_info_df = pd.DataFrame([{
                "Наименование клиента": additional_info["Наименование клиента"],
                "Период отчета": additional_info["Период отчета"],
                "Тарифный план": additional_info["Тарифный план"]
            }])
            for row in dataframe_to_rows(client_info_df, index=False, header=True):
                info_sla_worksheet.append(row)

            ### Заполнение информации о SLA, начиная с столбца "E"
            sla_info_df = pd.DataFrame(additional_info["SLA Информация"])
            for r_idx, row in enumerate(dataframe_to_rows(sla_info_df, index=False, header=True), start=1):
                for c_idx, value in enumerate(row, start=5):
                    info_sla_worksheet.cell(row=r_idx, column=c_idx, value=value)

            customize_combined_sheet(info_sla_worksheet)

            ### Установка ширины столбцов
            info_sla_worksheet.column_dimensions['A'].width = max(40, len(additional_info["Наименование клиента"]))
            info_sla_worksheet.column_dimensions['B'].width = max(15, len(additional_info["Период отчета"]))
            info_sla_worksheet.column_dimensions['C'].width = 16
            info_sla_worksheet.column_dimensions['E'].width = 15
            info_sla_worksheet.column_dimensions['F'].width = 28
            info_sla_worksheet.column_dimensions['G'].width = 29

        return str(file_path)
    except Exception as e:
        scripts_error_logger.error(f"Ошибка при создании Excel отчета: {traceback.format_exc()}")
        return None

def customize_excel_sheet(worksheet):
    """
    Настраивает формат первого листа с отчетом.

    Аргументы:
    worksheet (Worksheet): Лист Excel для настройки.
    """
    try:
        ### Настройки шрифтов, выравнивания и цветов
        header_font = Font(bold=True)
        content_font = Font(bold=False)
        header_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        wrap_text_alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

        header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        sla_violation_fill = PatternFill(start_color='FEF6BE', end_color='FEF6BE', fill_type='solid')

        ### Определение границ ячеек с серым цветом
        gray_border = Border(left=Side(style='thin', color='808080'),
                             right=Side(style='thin', color='808080'),
                             top=Side(style='thin', color='808080'),
                             bottom=Side(style='thin', color='808080'))

        ### Настройка заголовков таблицы
        for cell in worksheet["1:1"]:
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = gray_border

        ### Установка ширины столбцов
        worksheet.column_dimensions['A'].width = 12
        worksheet.column_dimensions['B'].width = 16
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 55
        worksheet.column_dimensions['E'].width = 25
        worksheet.column_dimensions['F'].width = 25
        worksheet.column_dimensions['G'].width = 12
        worksheet.column_dimensions['H'].width = 20
        worksheet.column_dimensions['I'].width = 20
        worksheet.column_dimensions['J'].width = 20
        worksheet.column_dimensions['K'].width = 20
        worksheet.column_dimensions['L'].width = 20

        ### Настройка формата ячеек и границ
        for row in worksheet.iter_rows(min_row=2, max_col=12, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = wrap_text_alignment
                cell.border = gray_border
                if cell.column_letter == 'L' and 'Да' in cell.value:
                    cell.fill = sla_violation_fill

        ### Добавление фильтров и закрепление заголовков
        worksheet.auto_filter.ref = "A1:L1"
        worksheet.freeze_panes = "A2"

    except Exception as e:
        scripts_error_logger.error(f"Ошибка при настройке первого листа Excel: {traceback.format_exc()}")

def customize_combined_sheet(worksheet):
    """
    Настраивает формат второго листа с информацией о клиенте и SLA.

    Аргументы:
    worksheet (Worksheet): Лист Excel для настройки.
    """
    try:
        ### Настройки шрифтов, выравнивания и цветов
        header_font = Font(bold=True)
        content_font = Font(bold=False)
        header_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        wrap_text_alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        gray_border = Border(left=Side(style='thin', color='808080'),
                             right=Side(style='thin', color='808080'),
                             top=Side(style='thin', color='808080'),
                             bottom=Side(style='thin', color='808080'))

        ### Настройка заголовков информации о клиенте
        for cell in worksheet[1]:
            if cell.column_letter in ['A', 'B', 'C']:
                cell.font = header_font
                cell.alignment = header_alignment
                cell.fill = header_fill
                cell.border = gray_border

        ### Настройка данных информации о клиенте
        for row in worksheet.iter_rows(min_row=2, max_row=2):
            for cell in row:
                if cell.column_letter in ['A', 'B', 'C']:
                    cell.font = content_font
                    cell.alignment = center_alignment if cell.column_letter in ['B', 'C'] else wrap_text_alignment
                    cell.border = gray_border

        ### Настройка заголовков информации о SLA
        for cell in worksheet[1]:
            if cell.column_letter in ['E', 'F', 'G']:
                cell.font = header_font
                cell.alignment = header_alignment
                cell.fill = header_fill
                cell.border = gray_border

        ### Настройка данных информации о SLA
        priority_fill_colors = {
            "Критический": PatternFill(start_color='F2DCDB', end_color='F2DCDB', fill_type='solid'),
            "Высокий": PatternFill(start_color='FDE9D9', end_color='FDE9D9', fill_type='solid'),
            "Средний": PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),
            "Низкий": PatternFill(start_color='E7FFE7', end_color='E7FFE7', fill_type='solid')
        }

        ### Настройка данных информации о SLA
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            priority = row[4].value
            fill_color = priority_fill_colors.get(priority, None)
            for cell in row:
                if cell.column_letter in ['E', 'F', 'G']:
                    cell.font = content_font
                    cell.alignment = center_alignment if cell.column_letter in ['F', 'G'] else wrap_text_alignment
                    cell.border = gray_border
                    if fill_color:
                        cell.fill = fill_color

    except Exception as e:
        scripts_error_logger.error(f"Ошибка при настройке второго листа Excel: {traceback.format_exc()}")

def generate_report(data, client_name, start_date, end_date):
    """
    Генерирует отчет с данными и отправляет его на email.

    Аргументы:
    data (QuerySet): Набор данных с информацией о тикетах.
    client_name (str): Название клиента.
    start_date (str): Дата начала отчетного периода.
    end_date (str): Дата окончания отчетного периода.

    Возвращает:
    str: Путь к созданному файлу отчета или None в случае ошибки.
    """
    try:
        scripts_info_logger.info(f"Начало генерации отчета для клиента: {client_name}")
        report_data = prepare_report_data(data)
        file_path = create_excel_report(report_data, client_name, start_date, end_date)
        if not file_path:
            raise ValueError("Не удалось создать отчет.")
        return file_path
    except Exception as e:
        scripts_error_logger.error(f"Ошибка при генерации отчета для {client_name}: {traceback.format_exc()}")
        return None

def send_report(recipient_email, file_path, client_name, start_date, end_date):
    """
    Отправляет отчет на email.

    Аргументы:
    recipient_email (str): Email получателя.
    file_path (str): Путь к файлу отчета.
    client_name (str): Название клиента.
    start_date (str): Дата начала отчетного периода.
    end_date (str): Дата окончания отчетного периода.

    Возвращает:
    bool: True, если отчет успешно отправлен, иначе False.
    """
    try:
        if not file_path or not os.path.exists(file_path):
            raise FileNotFoundError(f"Файл отчета не найден по пути: {file_path}")

        ### Чтение содержимого файла отчета
        with open(file_path, 'rb') as f:
            file_content = f.read()

        file_name = os.path.basename(file_path)

        ### Формирование и отправка email
        email = EmailMessage(
            subject=f'Отчет по заявкам для {client_name}',
            body=f'Добрый день!\nВо вложении отчёт по заявкам за период с {start_date} по {end_date}.',
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[recipient_email]
        )
        email.attach(file_name, file_content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        email.send()

        scripts_info_logger.info(f"Отчет успешно отправлен на {recipient_email}")
        return True
    except Exception as error_message:
        scripts_error_logger.error(f"Ошибка при отправке отчета на email {recipient_email}: {traceback.format_exc()}")
        return False
